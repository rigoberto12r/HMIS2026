"""
Report generation service with predefined templates and custom query execution.
Supports clinical, financial, and operational reports with multiple export formats.
"""

import csv
import io
import json
import uuid
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy import func, select, and_, or_, desc, asc
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.appointments.models import Appointment, Provider
from app.modules.billing.models import Invoice, Payment, InsuranceClaim, ChargeItem
from app.modules.emr.models import Encounter, Diagnosis
from app.modules.patients.models import Patient, PatientInsurance
from app.modules.pharmacy.models import Prescription, Product
from app.modules.reports.models import ReportDefinition, ReportExecution, ScheduledReport
from app.modules.reports.schemas import QueryConfig, QueryFilter


class ReportService:
    """Service for generating and managing reports."""

    def __init__(self, db: AsyncSession, tenant_id: str):
        self.db = db
        self.tenant_id = tenant_id
        self.export_dir = Path("exports/reports")
        self.export_dir.mkdir(parents=True, exist_ok=True)

    # ─── Predefined Report Templates ─────────────────────────────

    async def generate_patient_demographics_report(
        self, start_date: date | None = None, end_date: date | None = None
    ) -> dict:
        """
        Patient statistics by age group, gender, and insurance coverage.
        """
        query = select(
            Patient.gender,
            func.count(Patient.id).label("count"),
            func.avg(
                func.extract("year", func.now()) - func.extract("year", Patient.birth_date)
            ).label("avg_age"),
        ).where(
            Patient.tenant_id == self.tenant_id, Patient.is_active == True
        )

        if start_date:
            query = query.where(Patient.created_at >= start_date)
        if end_date:
            query = query.where(Patient.created_at <= end_date)

        query = query.group_by(Patient.gender)

        result = await self.db.execute(query)
        gender_stats = result.all()

        # Age distribution
        age_query = select(
            func.case(
                (
                    func.extract("year", func.now())
                    - func.extract("year", Patient.birth_date)
                    < 18,
                    "0-17",
                ),
                (
                    func.extract("year", func.now())
                    - func.extract("year", Patient.birth_date)
                    < 30,
                    "18-29",
                ),
                (
                    func.extract("year", func.now())
                    - func.extract("year", Patient.birth_date)
                    < 50,
                    "30-49",
                ),
                (
                    func.extract("year", func.now())
                    - func.extract("year", Patient.birth_date)
                    < 65,
                    "50-64",
                ),
                else_="65+",
            ).label("age_group"),
            func.count(Patient.id).label("count"),
        ).where(Patient.tenant_id == self.tenant_id, Patient.is_active == True)

        if start_date:
            age_query = age_query.where(Patient.created_at >= start_date)
        if end_date:
            age_query = age_query.where(Patient.created_at <= end_date)

        age_query = age_query.group_by("age_group")

        age_result = await self.db.execute(age_query)
        age_distribution = age_result.all()

        # Insurance coverage
        insurance_query = select(
            PatientInsurance.insurer_name, func.count(PatientInsurance.id).label("count")
        ).where(
            PatientInsurance.tenant_id == self.tenant_id,
            PatientInsurance.is_active == True,
            PatientInsurance.status == "active",
        )

        if start_date:
            insurance_query = insurance_query.where(PatientInsurance.created_at >= start_date)
        if end_date:
            insurance_query = insurance_query.where(PatientInsurance.created_at <= end_date)

        insurance_query = insurance_query.group_by(PatientInsurance.insurer_name).order_by(
            desc("count")
        )

        insurance_result = await self.db.execute(insurance_query)
        insurance_stats = insurance_result.all()

        return {
            "gender_distribution": [
                {"gender": row.gender, "count": row.count, "avg_age": float(row.avg_age or 0)}
                for row in gender_stats
            ],
            "age_distribution": [
                {"age_group": row.age_group, "count": row.count} for row in age_distribution
            ],
            "insurance_coverage": [
                {"insurer": row.insurer_name, "patient_count": row.count}
                for row in insurance_stats
            ],
        }

    async def generate_diagnosis_trends_report(
        self, start_date: date | None = None, end_date: date | None = None, limit: int = 20
    ) -> dict:
        """
        Top diagnoses over time period with frequency analysis.
        """
        query = (
            select(
                Diagnosis.icd10_code,
                Diagnosis.description,
                func.count(Diagnosis.id).label("count"),
                func.count(func.distinct(Diagnosis.patient_id)).label("unique_patients"),
            )
            .join(Encounter, Diagnosis.encounter_id == Encounter.id)
            .where(
                Diagnosis.tenant_id == self.tenant_id,
                Diagnosis.is_active == True,
                Encounter.is_active == True,
            )
        )

        if start_date:
            query = query.where(Encounter.start_datetime >= start_date)
        if end_date:
            query = query.where(Encounter.start_datetime <= end_date)

        query = (
            query.group_by(Diagnosis.icd10_code, Diagnosis.description)
            .order_by(desc("count"))
            .limit(limit)
        )

        result = await self.db.execute(query)
        diagnoses = result.all()

        return {
            "top_diagnoses": [
                {
                    "icd10_code": row.icd10_code,
                    "description": row.description,
                    "total_occurrences": row.count,
                    "unique_patients": row.unique_patients,
                }
                for row in diagnoses
            ]
        }

    async def generate_provider_productivity_report(
        self, start_date: date | None = None, end_date: date | None = None
    ) -> dict:
        """
        Provider statistics: appointments, encounters, and productivity metrics.
        """
        query = (
            select(
                Provider.id,
                Provider.first_name,
                Provider.last_name,
                Provider.specialty_name,
                func.count(func.distinct(Appointment.id)).label("total_appointments"),
                func.count(func.distinct(Encounter.id)).label("total_encounters"),
                func.count(
                    func.distinct(
                        func.case((Appointment.status == "completed", Appointment.id), else_=None)
                    )
                ).label("completed_appointments"),
                func.count(
                    func.distinct(
                        func.case((Appointment.status == "no_show", Appointment.id), else_=None)
                    )
                ).label("no_shows"),
            )
            .outerjoin(Appointment, Provider.id == Appointment.provider_id)
            .outerjoin(Encounter, Provider.id == Encounter.provider_id)
            .where(Provider.tenant_id == self.tenant_id, Provider.is_active == True)
        )

        if start_date:
            query = query.where(
                or_(
                    Appointment.scheduled_start >= start_date,
                    Encounter.start_datetime >= start_date,
                )
            )
        if end_date:
            query = query.where(
                or_(Appointment.scheduled_start <= end_date, Encounter.start_datetime <= end_date)
            )

        query = query.group_by(
            Provider.id, Provider.first_name, Provider.last_name, Provider.specialty_name
        ).order_by(desc("total_appointments"))

        result = await self.db.execute(query)
        providers = result.all()

        return {
            "provider_productivity": [
                {
                    "provider_id": str(row.id),
                    "name": f"{row.first_name} {row.last_name}",
                    "specialty": row.specialty_name,
                    "total_appointments": row.total_appointments,
                    "completed_appointments": row.completed_appointments,
                    "no_shows": row.no_shows,
                    "total_encounters": row.total_encounters,
                    "completion_rate": (
                        round(row.completed_appointments / row.total_appointments * 100, 2)
                        if row.total_appointments > 0
                        else 0
                    ),
                }
                for row in providers
            ]
        }

    async def generate_revenue_analysis_report(
        self, start_date: date | None = None, end_date: date | None = None
    ) -> dict:
        """
        Revenue analysis by service category, payment method, and time period.
        """
        # Revenue by invoice status
        invoice_query = select(
            Invoice.status,
            func.count(Invoice.id).label("count"),
            func.sum(Invoice.grand_total).label("total_amount"),
        ).where(Invoice.tenant_id == self.tenant_id, Invoice.is_active == True)

        if start_date:
            invoice_query = invoice_query.where(Invoice.created_at >= start_date)
        if end_date:
            invoice_query = invoice_query.where(Invoice.created_at <= end_date)

        invoice_query = invoice_query.group_by(Invoice.status)

        invoice_result = await self.db.execute(invoice_query)
        invoice_stats = invoice_result.all()

        # Payments by method
        payment_query = select(
            Payment.payment_method,
            func.count(Payment.id).label("count"),
            func.sum(Payment.amount).label("total_amount"),
        ).where(Payment.tenant_id == self.tenant_id, Payment.is_active == True)

        if start_date:
            payment_query = payment_query.where(Payment.received_at >= start_date)
        if end_date:
            payment_query = payment_query.where(Payment.received_at <= end_date)

        payment_query = payment_query.group_by(Payment.payment_method)

        payment_result = await self.db.execute(payment_query)
        payment_stats = payment_result.all()

        # Revenue by charge item category
        charge_query = (
            select(
                ChargeItem.description, func.sum(ChargeItem.total).label("total_amount")
            )
            .where(
                ChargeItem.tenant_id == self.tenant_id,
                ChargeItem.is_active == True,
                ChargeItem.status.in_(["invoiced", "paid"]),
            )
            .group_by(ChargeItem.description)
            .order_by(desc("total_amount"))
            .limit(10)
        )

        if start_date:
            charge_query = charge_query.where(ChargeItem.created_at >= start_date)
        if end_date:
            charge_query = charge_query.where(ChargeItem.created_at <= end_date)

        charge_result = await self.db.execute(charge_query)
        charge_stats = charge_result.all()

        return {
            "invoice_summary": [
                {
                    "status": row.status,
                    "count": row.count,
                    "total_amount": float(row.total_amount or 0),
                }
                for row in invoice_stats
            ],
            "payment_methods": [
                {
                    "method": row.payment_method,
                    "count": row.count,
                    "total_amount": float(row.total_amount or 0),
                }
                for row in payment_stats
            ],
            "top_services": [
                {"service": row.description, "total_revenue": float(row.total_amount or 0)}
                for row in charge_stats
            ],
        }

    async def generate_insurance_claims_report(
        self, start_date: date | None = None, end_date: date | None = None
    ) -> dict:
        """
        Insurance claims by status, insurer, and amounts.
        """
        query = (
            select(
                InsuranceClaim.insurer_name,
                InsuranceClaim.status,
                func.count(InsuranceClaim.id).label("count"),
                func.sum(InsuranceClaim.total_claimed).label("total_claimed"),
                func.sum(InsuranceClaim.total_approved).label("total_approved"),
                func.sum(InsuranceClaim.total_denied).label("total_denied"),
            )
            .where(InsuranceClaim.tenant_id == self.tenant_id, InsuranceClaim.is_active == True)
            .group_by(InsuranceClaim.insurer_name, InsuranceClaim.status)
            .order_by(InsuranceClaim.insurer_name)
        )

        if start_date:
            query = query.where(InsuranceClaim.created_at >= start_date)
        if end_date:
            query = query.where(InsuranceClaim.created_at <= end_date)

        result = await self.db.execute(query)
        claims = result.all()

        return {
            "claims_summary": [
                {
                    "insurer": row.insurer_name,
                    "status": row.status,
                    "count": row.count,
                    "total_claimed": float(row.total_claimed or 0),
                    "total_approved": float(row.total_approved or 0),
                    "total_denied": float(row.total_denied or 0),
                    "approval_rate": (
                        round(
                            (row.total_approved or 0) / (row.total_claimed or 1) * 100,
                            2,
                        )
                    ),
                }
                for row in claims
            ]
        }

    async def generate_appointment_statistics_report(
        self, start_date: date | None = None, end_date: date | None = None
    ) -> dict:
        """
        Appointment statistics by status, type, and specialty.
        """
        # By status
        status_query = select(
            Appointment.status, func.count(Appointment.id).label("count")
        ).where(Appointment.tenant_id == self.tenant_id, Appointment.is_active == True)

        if start_date:
            status_query = status_query.where(Appointment.scheduled_start >= start_date)
        if end_date:
            status_query = status_query.where(Appointment.scheduled_start <= end_date)

        status_query = status_query.group_by(Appointment.status)

        status_result = await self.db.execute(status_query)
        status_stats = status_result.all()

        # By type
        type_query = select(
            Appointment.appointment_type, func.count(Appointment.id).label("count")
        ).where(Appointment.tenant_id == self.tenant_id, Appointment.is_active == True)

        if start_date:
            type_query = type_query.where(Appointment.scheduled_start >= start_date)
        if end_date:
            type_query = type_query.where(Appointment.scheduled_start <= end_date)

        type_query = type_query.group_by(Appointment.appointment_type)

        type_result = await self.db.execute(type_query)
        type_stats = type_result.all()

        # By specialty
        specialty_query = (
            select(Provider.specialty_name, func.count(Appointment.id).label("count"))
            .join(Provider, Appointment.provider_id == Provider.id)
            .where(
                Appointment.tenant_id == self.tenant_id,
                Appointment.is_active == True,
                Provider.is_active == True,
            )
            .group_by(Provider.specialty_name)
            .order_by(desc("count"))
        )

        if start_date:
            specialty_query = specialty_query.where(Appointment.scheduled_start >= start_date)
        if end_date:
            specialty_query = specialty_query.where(Appointment.scheduled_start <= end_date)

        specialty_result = await self.db.execute(specialty_query)
        specialty_stats = specialty_result.all()

        return {
            "by_status": [
                {"status": row.status, "count": row.count} for row in status_stats
            ],
            "by_type": [
                {"type": row.appointment_type, "count": row.count} for row in type_stats
            ],
            "by_specialty": [
                {"specialty": row.specialty_name or "N/A", "count": row.count}
                for row in specialty_stats
            ],
        }

    # ─── Custom Report Execution ─────────────────────────────────

    async def execute_custom_report(
        self, query_config: QueryConfig, executed_by: uuid.UUID | None = None
    ) -> tuple[list[dict], list[str]]:
        """
        Execute a custom report based on query configuration.
        Returns tuple of (data, columns).
        """
        # Map data source to model
        model_map = {
            "patients": Patient,
            "appointments": Appointment,
            "billing": Invoice,
            "pharmacy": Prescription,
            "emr": Encounter,
        }

        model = model_map.get(query_config.data_source)
        if not model:
            raise ValueError(f"Invalid data source: {query_config.data_source}")

        # Build query
        query = select(model).where(
            model.tenant_id == self.tenant_id, model.is_active == True
        )

        # Apply filters
        for filter_spec in query_config.filters:
            query = self._apply_filter(query, model, filter_spec)

        # Apply sorting
        for sort_spec in query_config.sort:
            field = getattr(model, sort_spec.field, None)
            if field is not None:
                query = (
                    query.order_by(desc(field))
                    if sort_spec.direction == "desc"
                    else query.order_by(asc(field))
                )

        # Apply limit
        if query_config.limit:
            query = query.limit(query_config.limit)

        result = await self.db.execute(query)
        rows = result.scalars().all()

        # Extract data
        data = []
        columns = query_config.fields if query_config.fields else []

        for row in rows:
            row_data = {}
            if columns:
                for field in columns:
                    value = getattr(row, field, None)
                    # Convert special types to JSON-serializable
                    if isinstance(value, (datetime, date)):
                        value = value.isoformat()
                    elif isinstance(value, uuid.UUID):
                        value = str(value)
                    row_data[field] = value
            else:
                # Return all fields if none specified
                for col in row.__table__.columns:
                    value = getattr(row, col.name, None)
                    if isinstance(value, (datetime, date)):
                        value = value.isoformat()
                    elif isinstance(value, uuid.UUID):
                        value = str(value)
                    row_data[col.name] = value
                if not columns:
                    columns = [col.name for col in row.__table__.columns]

        data.append(row_data)

        return data, columns

    def _apply_filter(self, query, model, filter_spec: QueryFilter):
        """Apply a single filter to the query."""
        field = getattr(model, filter_spec.field, None)
        if field is None:
            return query

        if filter_spec.operator == "equals":
            return query.where(field == filter_spec.value)
        elif filter_spec.operator == "not_equals":
            return query.where(field != filter_spec.value)
        elif filter_spec.operator == "contains":
            return query.where(field.ilike(f"%{filter_spec.value}%"))
        elif filter_spec.operator == "gt":
            return query.where(field > filter_spec.value)
        elif filter_spec.operator == "gte":
            return query.where(field >= filter_spec.value)
        elif filter_spec.operator == "lt":
            return query.where(field < filter_spec.value)
        elif filter_spec.operator == "lte":
            return query.where(field <= filter_spec.value)
        elif filter_spec.operator == "in":
            return query.where(field.in_(filter_spec.value))
        elif filter_spec.operator == "between" and isinstance(filter_spec.value, list):
            return query.where(field.between(filter_spec.value[0], filter_spec.value[1]))

        return query

    # ─── Export Functions ────────────────────────────────────────

    async def export_to_csv(
        self, data: list[dict], columns: list[str], filename: str
    ) -> tuple[str, int]:
        """
        Export data to CSV format.
        Returns tuple of (file_path, file_size).
        """
        filepath = self.export_dir / f"{filename}.csv"

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            if not data:
                return str(filepath), 0

            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(data)

        file_size = filepath.stat().st_size
        return str(filepath), file_size

    async def export_to_excel(
        self, data: list[dict], columns: list[str], filename: str, report_title: str = ""
    ) -> tuple[str, int]:
        """
        Export data to Excel format with formatting.
        Returns tuple of (file_path, file_size).
        """
        filepath = self.export_dir / f"{filename}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Add title
        if report_title:
            ws.merge_cells("A1:{}1".format(chr(65 + len(columns) - 1)))
            title_cell = ws["A1"]
            title_cell.value = report_title
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal="center")
            row_offset = 3
        else:
            row_offset = 1

        # Add headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=row_offset, column=col_idx)
            cell.value = column.replace("_", " ").title()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for row_idx, row_data in enumerate(data, row_offset + 1):
            for col_idx, column in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(column, ""))

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        file_size = filepath.stat().st_size
        return str(filepath), file_size

    async def export_to_pdf(
        self,
        data: list[dict],
        columns: list[str],
        filename: str,
        report_title: str = "Report",
        report_subtitle: str = "",
    ) -> tuple[str, int]:
        """
        Export data to PDF format with table layout.
        Returns tuple of (file_path, file_size).
        """
        filepath = self.export_dir / f"{filename}.pdf"

        doc = SimpleDocTemplate(str(filepath), pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Heading1"], fontSize=18, spaceAfter=12
        )
        elements.append(Paragraph(report_title, title_style))

        if report_subtitle:
            subtitle_style = ParagraphStyle(
                "CustomSubtitle", parent=styles["Normal"], fontSize=10, spaceAfter=20
            )
            elements.append(Paragraph(report_subtitle, subtitle_style))

        # Prepare table data
        table_data = []

        # Headers
        table_data.append([col.replace("_", " ").title() for col in columns])

        # Data rows
        for row in data[:100]:  # Limit to first 100 rows for PDF
            table_data.append([str(row.get(col, "")) for col in columns])

        # Create table
        if table_data:
            col_widths = [doc.width / len(columns)] * len(columns)
            table = Table(table_data, colWidths=col_widths, repeatRows=1)

            # Style
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ]
                )
            )

            elements.append(table)

        # Build PDF
        doc.build(elements)
        file_size = filepath.stat().st_size
        return str(filepath), file_size

    # ─── Report Definition Management ────────────────────────────

    async def create_report_definition(
        self, definition_data: dict, created_by: uuid.UUID | None = None
    ) -> ReportDefinition:
        """Create a new report definition."""
        definition = ReportDefinition(
            **definition_data, tenant_id=self.tenant_id, created_by=created_by
        )
        self.db.add(definition)
        await self.db.commit()
        await self.db.refresh(definition)
        return definition

    async def get_report_definition(self, definition_id: uuid.UUID) -> ReportDefinition | None:
        """Get report definition by ID."""
        query = select(ReportDefinition).where(
            ReportDefinition.id == definition_id,
            ReportDefinition.tenant_id == self.tenant_id,
            ReportDefinition.is_active == True,
        )
        result = await self.db.execute(query)
        return result.scalar_one_or_none()

    async def list_report_definitions(
        self, report_type: str | None = None, skip: int = 0, limit: int = 100
    ) -> list[ReportDefinition]:
        """List report definitions with optional filtering."""
        query = select(ReportDefinition).where(
            ReportDefinition.tenant_id == self.tenant_id, ReportDefinition.is_active == True
        )

        if report_type:
            query = query.where(ReportDefinition.report_type == report_type)

        query = query.order_by(desc(ReportDefinition.created_at)).offset(skip).limit(limit)

        result = await self.db.execute(query)
        return list(result.scalars().all())

    async def create_report_execution(
        self,
        definition_id: uuid.UUID | None,
        executed_by: uuid.UUID | None,
        parameters: dict | None,
    ) -> ReportExecution:
        """Create a report execution record."""
        execution = ReportExecution(
            tenant_id=self.tenant_id,
            report_definition_id=definition_id,
            executed_by=executed_by,
            parameters=parameters,
            status="pending",
        )
        self.db.add(execution)
        await self.db.commit()
        await self.db.refresh(execution)
        return execution

    async def update_report_execution(
        self,
        execution_id: uuid.UUID,
        status: str,
        row_count: int | None = None,
        execution_time_ms: int | None = None,
        file_path: str | None = None,
        file_format: str | None = None,
        file_size_bytes: int | None = None,
        error_message: str | None = None,
        result_data: dict | None = None,
    ) -> None:
        """Update report execution with results."""
        query = (
            select(ReportExecution)
            .where(
                ReportExecution.id == execution_id,
                ReportExecution.tenant_id == self.tenant_id,
            )
        )
        result = await self.db.execute(query)
        execution = result.scalar_one_or_none()

        if execution:
            execution.status = status
            if row_count is not None:
                execution.row_count = row_count
            if execution_time_ms is not None:
                execution.execution_time_ms = execution_time_ms
            if file_path:
                execution.file_path = file_path
            if file_format:
                execution.file_format = file_format
            if file_size_bytes is not None:
                execution.file_size_bytes = file_size_bytes
            if error_message:
                execution.error_message = error_message
            if result_data:
                execution.result_data = result_data

            await self.db.commit()
